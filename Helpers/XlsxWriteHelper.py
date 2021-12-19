from typing import Sequence
from openpyxl import Workbook, load_workbook
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.chart.legend import LegendEntry
from openpyxl.chart.pie_chart import PieChart
from openpyxl.chart.reference import Reference
from openpyxl.drawing.fill import PatternFillProperties
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, alignment
from Enums.GuidParameter import GuidParameter
from openpyxl.chart.bar_chart import BarChart, GraphicalProperties

class XlsxWriteHelper:
    def __init__(self, file, exist=True):
        self.file_path = file
        if (exist):
            self.workbook = load_workbook(filename=file)
        else:
            self.workbook = Workbook()

        
        self.active_worksheet=None

        self.middle_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left')
        self.right_align = Alignment(horizontal='right')

    def add_and_set_worksheet(self,worksheet_name):
        created_worksheet = self.workbook.create_sheet(title=worksheet_name, index=0)
        self.workbook.active = 0
        self.active_worksheet = self.workbook.active 

    def close_workbook(self):
        self.workbook.save(self.file_path)
        self.workbook.close()

    def close_workbook_without_save(self):
        self.workbook.close()

    def write_data_to_worksheet_no_color(self, data, time, start_row):
        if(self.active_worksheet == None):
            raise ValueError("Active worksheet is none.")

        row=start_row
        row = self.write_data_to_worksheet_with_color(data, time, start_row)

        return row

    def write_data_to_worksheet_with_color(self, data, time, start_row, hightlight_fill=None, fill=None):
        if(self.active_worksheet == None):
            raise ValueError("Active worksheet is none.")

        row=start_row
        for key in data:
            if (key=="Total"):
                continue

            if (row-start_row<5):
                self.create_normal_left_align_cell(row, 2, key, hightlight_fill)
                self.create_normal_right_align_cell(row, 3, data[key], hightlight_fill)
            else:
                self.create_normal_left_align_cell(row, 2, key, fill)
                self.create_normal_right_align_cell(row, 3, data[key], fill)

            row += 1

        self.create_normal_left_align_cell(row, 2, "Total", fill)
        self.create_normal_right_align_cell(row, 3, data["Total"], fill)

        self.create_merge_cell(start_row, row, 1, 1, time, fill)
        return row


    def write_data_to_worksheet_hightlight_orangebrown_top_5(self, data, time, start_row):
        if(self.active_worksheet == None):
            raise ValueError("Active worksheet is none.")

        hightlight_fill = PatternFill('solid',start_color=GuidParameter.ORANGE_BROWN, end_color=GuidParameter.ORANGE_BROWN)
        row = self.write_data_to_worksheet_with_color(data, time, start_row, hightlight_fill)

        return row

    def write_data_to_worksheet_hightlight_yellow_top_5_and_fill_green(self, data, time, start_row):
        if(self.active_worksheet == None):
            raise ValueError("Active worksheet is none.")

        hightlight_fill = PatternFill('solid', start_color=GuidParameter.YELLOW, end_color=GuidParameter.YELLOW)
        normal_fill = PatternFill('solid', start_color=GuidParameter.GREEN, end_color=GuidParameter.GREEN)
        row = self.write_data_to_worksheet_with_color(data, time, start_row, hightlight_fill, normal_fill)

        return row

    def write_data_to_worksheet_top_10_no_color(self, data, time, start_row):
        if(self.active_worksheet == None):
            raise ValueError("Active worksheet is none.")

        tmp = dict()
        count=0
        for key in data:
            tmp[key] = data[key]
            count+=1
            if (count==11):
                break

        tmp["Total"] = data["Total"]
        self.write_data_to_worksheet_no_color(tmp, time, start_row)

    def write_support_ticket_data_to_worksheet(self, data, start_row):
        # tech and their tickets
        row=start_row
        for key in data:
            if (not key=="Total"):
                self.create_normal_left_align_cell(row, 1, key)
                self.create_normal_right_align_cell(row, 2, data[key])
                row+=1
        
        self.create_normal_left_align_cell(row, 1, "Total")
        self.create_normal_right_align_cell(row, 2, data["Total"])

        row = row + 2
        tech_layer_start_row = row

        # layer tech
        # Layer 1 Techs: CTL Student Workers 
        student_worker = data["CSTL Student Worker"] + data["CSTL 2 Student Worker"]
        self.create_normal_left_align_cell(row, 1, "Layer 1 Techs: CTL Student Workers ")
        self.create_normal_right_align_cell(row, 2, student_worker)
        row+=1

        # Layer 2 Techs: CTL Graduate Assistants 
        ga = data["CSTL GA"]
        self.create_normal_left_align_cell(row, 1, "Layer 2 Techs: CTL Graduate Assistants ")
        self.create_normal_right_align_cell(row, 2, ga)
        row+=1

        # Layer 3 Techs: CTL Administrators 
        ctl_admins = data["Kris Baranovic"] + data["Mary Harriet Talbut"]
        self.create_normal_left_align_cell(row, 1, "Layer 3 Techs: CTL Administrators ")
        self.create_normal_right_align_cell(row, 2, ctl_admins)
        row+=1

        # Layer 4 Techs: IT Administration 
        it_Admins = data["Total"] - student_worker - ga - ctl_admins
        self.create_normal_left_align_cell(row, 1, "Layer 4 Techs: IT Administration ")
        self.create_normal_right_align_cell(row, 2, it_Admins)

        return tech_layer_start_row


    def create_normal_left_align_cell(self, row, column, value, fill=None):
        cell = self.active_worksheet.cell(row=row, column=column)
        cell.value = value
        cell.alignment = self.left_align
        if fill:
            cell.fill = fill

        return cell

    def create_normal_right_align_cell(self, row, column, value, fill=None):
        cell = self.active_worksheet.cell(row=row, column=column)
        cell.value = value
        cell.alignment = self.right_align
        if fill:
            cell.fill = fill

        return cell

    def create_merge_cell(self, start_row, end_row, start_col, end_col, value, fill=None):
        self.active_worksheet.merge_cells(start_row=start_row, end_row=end_row, start_column=start_col, end_column=end_col)
        time_cell = self.active_worksheet.cell(row=start_row, column=start_col)
        time_cell.value = value
        time_cell.alignment = self.middle_align
        if fill:
            time_cell.fill = fill

    def create_horizontal_bar_chart(self, title, data_start_col, data_end_col, data_start_row, data_end_row, cats_col, cats_start_row, cats_end_row, anchor):
        if (self.active_worksheet == None):
            raise ValueError("Active worksheet is none.")

        bar_chart = BarChart()
        bar_chart.type = "bar"
        bar_chart.style = 1
        bar_chart.title = title
        bar_chart.y_axis.title = "Tickets"
        bar_chart.x_axis.title = "Type"

        data_reference = Reference(self.active_worksheet, min_col=data_start_col, max_col=data_end_col, min_row=data_start_row, max_row=data_end_row)
        bar_chart.add_data(data_reference)

        cats_reference = Reference(self.active_worksheet, min_col=cats_col, min_row=cats_start_row, max_row=cats_end_row)
        bar_chart.set_categories(cats_reference)
    
        #style chart
        series = bar_chart.series[0]
        series.graphicalProperties.line.solidFill = GuidParameter.BLUE_FOR_BAR
        series.graphicalProperties.solidFill = GuidParameter.BLUE_FOR_BAR

        bar_chart.plot_area.graphicalProperties = GraphicalProperties(solidFill=GuidParameter.LIGHT_BLUE_FOR_BG)
        
        bar_chart.legend.legendEntry=[LegendEntry(idx=0, delete=True)] 

        chart_area_bg = GraphicalProperties(solidFill=GuidParameter.LIGHT_BLUE_FOR_BG)
        bar_chart.graphical_properties = chart_area_bg

        self.active_worksheet.add_chart(bar_chart, anchor=anchor)

    def create_pie_chart(self, title, data_col, data_start_row, data_end_row, cats_col, cats_start_row, cats_end_row, chart_anchor):
        piechart = PieChart()
        piechart.title = title
        
        data_reference = Reference(self.active_worksheet, min_col=data_col, min_row=data_start_row, max_row=data_end_row)
        piechart.add_data(data_reference)

        cats_reference = Reference(self.active_worksheet, min_col=cats_col, min_row=cats_start_row, max_row=cats_end_row)
        piechart.set_categories(cats_reference)

        # style
        piechart.style=1
        piechart.dataLabels = DataLabelList([DataLabel()])
        piechart.dataLabels.showCatName = True
        piechart.dataLabels.showLeaderLines = True

        self.active_worksheet.add_chart(piechart, chart_anchor)
        
        
